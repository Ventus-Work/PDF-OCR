"""Analyze ps-docparser output folders and write a compact QA report."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import sys

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from validators.output_quality import validate_bom_table, validate_table_contract


def _load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _is_compare_artifact(path: Path) -> bool:
    return "_compare" in path.parts


def _iter_tables(data):
    sections = data if isinstance(data, list) else data.get("sections", [])
    for section in sections:
        for table in section.get("tables", []) or []:
            yield section, table


def _non_empty_value_count(rows) -> int:
    count = 0
    for row in rows or []:
        if isinstance(row, dict):
            values = row.values()
        elif isinstance(row, (list, tuple)):
            values = row
        else:
            values = [row]
        count += sum(1 for value in values if str(value or "").strip())
    return count


def _xlsx_data_value_count(xlsx_path: Path, sheet_name: str, min_row: int) -> int:
    if not xlsx_path.exists():
        return 0
    try:
        workbook = load_workbook(xlsx_path, data_only=True)
    except Exception:
        return 0
    if sheet_name not in workbook.sheetnames:
        return 0

    worksheet = workbook[sheet_name]
    count = 0
    for row in worksheet.iter_rows(min_row=min_row, values_only=True):
        count += sum(1 for value in row if str(value or "").strip())
    return count


def _excel_value_loss_warnings(json_path: Path, data) -> list[str]:
    detail_value_count = 0
    for _, table in _iter_tables(data):
        if table.get("domain") == "estimate" and table.get("role") == "detail_table":
            detail_value_count += _non_empty_value_count(table.get("rows", []) or [])

    if detail_value_count < 10:
        return []

    excel_value_count = _xlsx_data_value_count(json_path.with_suffix(".xlsx"), "내역서", 4)
    if excel_value_count < max(3, detail_value_count // 10):
        return ["excel_value_loss_suspected"]
    return []


def _read_manifest_summary(manifest_path: Path) -> dict:
    if not manifest_path.exists():
        return {
            "inputs": 0,
            "representative": 0,
            "diagnostic": 0,
            "domains": {},
        }

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {
            "inputs": 0,
            "representative": 0,
            "diagnostic": 0,
            "domains": {},
            "error": "manifest_parse_error",
        }

    domains: dict[str, int] = {}
    representative = 0
    diagnostic = 0
    for entry in manifest.get("inputs", []) or []:
        primary = entry.get("primary") or {}
        if primary.get("role") == "representative":
            representative += 1
            domain = str(primary.get("domain") or "unknown")
            domains[domain] = domains.get(domain, 0) + 1
        for diagnostic_artifact in entry.get("diagnostics", []) or []:
            diagnostic += 1
            if diagnostic_artifact.get("role") == "representative":
                representative += 1
                domain = str(diagnostic_artifact.get("domain") or "unknown")
                domains[domain] = domains.get(domain, 0) + 1

    return {
        "inputs": len(manifest.get("inputs", []) or []),
        "representative": representative,
        "diagnostic": diagnostic,
        "domains": domains,
    }


def analyze_output_dir(output_dir: Path) -> dict:
    output_dir = Path(output_dir)
    json_files = sorted(
        path for path in output_dir.rglob("*.json") if not _is_compare_artifact(path)
    )
    xlsx_files = sorted(
        path for path in output_dir.rglob("*.xlsx") if not _is_compare_artifact(path)
    )
    manifest = output_dir / "RUN_MANIFEST.json"
    manifest_summary = _read_manifest_summary(manifest)

    mismatch_count = 0
    bad_header_count = 0
    quality_warnings: dict[str, int] = {}
    files = []

    for json_path in json_files:
        if json_path.name == "RUN_MANIFEST.json":
            continue
        try:
            data = _load_json(json_path)
        except Exception as exc:
            files.append({"file": str(json_path.relative_to(output_dir)), "error": str(exc)})
            continue

        table_count = 0
        row_count = 0
        for _, table in _iter_tables(data):
            table_count += 1
            headers = [str(header) for header in table.get("headers", [])]
            rows = table.get("rows", []) or []
            row_count += len(rows)

            for row in rows:
                if isinstance(row, dict) and list(row.keys()) != headers:
                    mismatch_count += 1

            saved_quality = table.get("quality")
            computed_quality = (
                validate_bom_table(headers, rows)
                if table.get("domain") == "bom" or table.get("type", "").startswith("BOM")
                else validate_table_contract(
                    headers,
                    rows,
                    domain=str(table.get("domain") or "generic"),
                    role=table.get("role"),
                )
            )

            warning_names: list[str] = []
            if isinstance(saved_quality, dict):
                warning_names.extend(saved_quality.get("warnings", []) or [])
            warning_names.extend(computed_quality.get("warnings", []) or [])

            for warning in dict.fromkeys(warning_names):
                quality_warnings[warning] = quality_warnings.get(warning, 0) + 1
                if warning == "self_repeating_composite_header":
                    bad_header_count += 1

        for warning in _excel_value_loss_warnings(json_path, data):
            quality_warnings[warning] = quality_warnings.get(warning, 0) + 1

        files.append({
            "file": str(json_path.relative_to(output_dir)),
            "tables": table_count,
            "rows": row_count,
        })

    return {
        "output_dir": str(output_dir),
        "json_files": len([p for p in json_files if p.name != "RUN_MANIFEST.json"]),
        "xlsx_files": len(xlsx_files),
        "has_manifest": manifest.exists(),
        "manifest": manifest_summary,
        "mismatch_count": mismatch_count,
        "bad_header_count": bad_header_count,
        "quality_warnings": quality_warnings,
        "files": files,
    }


def write_report(output_dir: Path, summary: dict) -> Path:
    status = "PASS"
    if summary["mismatch_count"] or summary["bad_header_count"]:
        status = "FAIL"
    elif summary["quality_warnings"]:
        status = "WARN"

    lines = [
        "# Output QA Report",
        "",
        f"- Status: `{status}`",
        f"- JSON files: `{summary['json_files']}`",
        f"- Excel files: `{summary['xlsx_files']}`",
        f"- RUN_MANIFEST.json: `{'yes' if summary['has_manifest'] else 'no'}`",
        f"- Manifest inputs: `{summary['manifest']['inputs']}`",
        f"- Manifest representative: `{summary['manifest']['representative']}`",
        f"- Manifest diagnostic: `{summary['manifest']['diagnostic']}`",
        f"- Header/key mismatch: `{summary['mismatch_count']}`",
        f"- Bad composite headers: `{summary['bad_header_count']}`",
        "",
        "## Quality Warnings",
        "",
    ]
    if summary["quality_warnings"]:
        for warning, count in sorted(summary["quality_warnings"].items()):
            lines.append(f"- `{warning}`: {count}")
    else:
        lines.append("- none")

    lines.extend(["", "## Manifest Domains", ""])
    if summary["manifest"]["domains"]:
        for domain, count in sorted(summary["manifest"]["domains"].items()):
            lines.append(f"- `{domain}`: {count}")
    else:
        lines.append("- none")

    lines.extend(["", "## Files", "", "| file | tables | rows |", "|---|---:|---:|"])
    for item in summary["files"]:
        if "error" in item:
            lines.append(f"| {item['file']} | error | {item['error']} |")
        else:
            lines.append(f"| {item['file']} | {item['tables']} | {item['rows']} |")

    report_path = Path(output_dir) / "OUTPUT_QA_REPORT.md"
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")
    return report_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Analyze ps-docparser output folder.")
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--json", action="store_true", help="Print machine-readable summary.")
    args = parser.parse_args()

    summary = analyze_output_dir(args.output_dir)
    report = write_report(args.output_dir, summary)
    if args.json:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print(f"Report: {report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
