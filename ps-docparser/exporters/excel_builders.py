"""
exporters/excel_builders.py — Excel 시트 빌더 함수 모음

Why: Phase 12 Step 12-2 분해 결과물.
     견적서/내역서/조건/범용 4종 시트 빌더와 내역서 헤더 그룹 상수를 담당.
     스타일은 excel_styles, 분류는 excel_classifier에서 import.

원본: exporters/excel_exporter.py L219~555 (4개 함수 + 1개 상수)
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter

from exporters.excel_classifier import _is_number, _row_style, _try_parse_number
from parsers.header_utils import normalize_header_text
from exporters.excel_styles import (
    _ALIGN_CENTER,
    _ALIGN_LEFT,
    _ALIGN_RIGHT,
    _BORDER_ALL,
    _BORDER_HEADER,
    _FILL_HEADER,
    _FILL_SECTION,
    _FILL_SUBTOTAL,
    _FILL_TITLE,
    _FONT_BODY,
    _FONT_HEADER,
    _FONT_NOTE,
    _FONT_SECTION,
    _FONT_SUBTOTAL,
    _FONT_TITLE,
    _apply_style,
)


# ═══════════════════════════════════════════════════════
# 내역서 복합 헤더 그룹 정의
# ═══════════════════════════════════════════════════════

# (그룹명, 포함 서브헤더 키워드 목록)
_DETAIL_HEADER_GROUPS = [
    ("품명",   ["품명"]),
    ("규격",   ["규격"]),
    ("단위",   ["단위"]),
    ("수량",   ["수량"]),
    ("재료비", ["재료비_단가", "재료비_금액"]),
    ("노무비", ["노무비_단가", "노무비_금액"]),
    ("경비",   ["경비_단가",   "경비_금액"]),
    ("합계",   ["합계_단가",   "합계_금액"]),
    ("비고",   ["비고"]),
]

_DETAIL_CANONICAL_KEYS = [
    sub_key
    for _, sub_keys in _DETAIL_HEADER_GROUPS
    for sub_key in sub_keys
]


def _canonicalize_detail_header_key(key: str) -> str:
    return normalize_header_text(str(key).strip())


def _build_detail_alias_map(headers: list[str], rows: list[dict]) -> dict[str, str]:
    alias_map: dict[str, str] = {}
    candidates: list[str] = list(headers)

    for row in rows[:5]:
        if isinstance(row, dict):
            candidates.extend(str(key) for key in row.keys())

    for candidate in candidates:
        canonical = _canonicalize_detail_header_key(candidate)
        if canonical in _DETAIL_CANONICAL_KEYS and canonical not in alias_map:
            alias_map[canonical] = str(candidate)

    return alias_map


def _get_detail_row_value(
    row: dict,
    canonical_key: str,
    alias_map: dict[str, str],
) -> str:
    actual_key = alias_map.get(canonical_key, canonical_key)
    if actual_key in row:
        return str(row.get(actual_key, "")).strip()
    if canonical_key in row:
        return str(row.get(canonical_key, "")).strip()
    return ""


# ═══════════════════════════════════════════════════════
# 견적서 시트 빌더
# ═══════════════════════════════════════════════════════

def _build_estimate_sheet(ws, table: dict, doc_title: str):
    """
    견적서 시트(요약 테이블) 작성.

    컬럼 레이아웃: NO | 명칭 | 규격 | 단위 | 수량 | 단가 | 금액 | 비고

    원본: excel_exporter.py L219~293
    """
    headers = table.get("headers", [])
    rows    = table.get("rows", [])

    if not headers:
        return

    # ── 문서 제목 행 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=doc_title or "견  적  서")
    _apply_style(title_cell, fill=_FILL_TITLE, font=_FONT_TITLE,
                 align=_ALIGN_CENTER, border=_BORDER_HEADER)
    ws.row_dimensions[1].height = 28

    # ── 헤더 행 ──
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
    ws.row_dimensions[2].height = 20

    # ── 데이터 행 ──
    first_key = headers[0]
    for row_idx, row in enumerate(rows, start=3):
        style = _row_style(row, first_key)
        ws.row_dimensions[row_idx].height = 16

        for col_idx, h in enumerate(headers, start=1):
            raw_val = str(row.get(h, "")).strip()
            cell = ws.cell(row=row_idx, column=col_idx)

            # [수정 D] 숫자 변환 적용: 금액·단가 컬럼 → int/float로 저장
            numeric = _try_parse_number(raw_val)
            if numeric is not None and col_idx >= 5:
                cell.value = numeric
                cell.number_format = '#,##0'
            else:
                cell.value = raw_val

            val = raw_val  # 스타일 판별용

            if style == "section":
                _apply_style(cell, fill=_FILL_SECTION, font=_FONT_SECTION,
                             align=_ALIGN_LEFT, border=_BORDER_ALL)
            elif style == "subtotal":
                _apply_style(cell, fill=_FILL_SUBTOTAL, font=_FONT_SUBTOTAL,
                             align=_ALIGN_RIGHT if (_is_number(val) or numeric is not None) else _ALIGN_LEFT,
                             border=_BORDER_ALL)
            else:
                if numeric is not None and col_idx >= 5:
                    _apply_style(cell, font=_FONT_BODY, align=_ALIGN_RIGHT, border=_BORDER_ALL)
                elif col_idx == 1:
                    _apply_style(cell, font=_FONT_BODY, align=_ALIGN_CENTER, border=_BORDER_ALL)
                else:
                    _apply_style(cell, font=_FONT_BODY, align=_ALIGN_LEFT, border=_BORDER_ALL)

    # ── 주석 행 ──
    notes = table.get("notes_in_table", [])
    if notes:
        note_row = len(rows) + 3
        note_cell = ws.cell(row=note_row, column=1,
                            value="※ " + " / ".join(notes))
        _apply_style(note_cell, font=_FONT_NOTE, align=_ALIGN_LEFT, border=None)
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=len(headers))

    # ── 컬럼 폭 설정 ──
    col_widths = {1: 6, 2: 28, 3: 14, 4: 6, 5: 7, 6: 14, 7: 14, 8: 12}
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_idx, 12)


# ═══════════════════════════════════════════════════════
# 내역서 시트 빌더
# ═══════════════════════════════════════════════════════

def _build_detail_sheet(ws, table: dict, doc_title: str):
    """
    내역서 시트(상세 테이블) 작성.

    복합 헤더 2행 구조:
        행1: 품명 | 규격 | 단위 | 수량 | 재료비(병합) | 노무비(병합) | 경비(병합) | 합계(병합) | 비고
        행2:  -   |  -   |  -   |  -   | 단가 | 금액  | 단가 | 금액 | 단가 | 금액 | 단가 | 금액 |  -

    원본: excel_exporter.py L315~448
    """
    headers = table.get("headers", [])
    rows    = table.get("rows", [])

    if not headers:
        return

    alias_map = _build_detail_alias_map(headers, rows)
    col_order = [sub_key for _, sub_keys in _DETAIL_HEADER_GROUPS for sub_key in sub_keys]
    active_groups = list(_DETAIL_HEADER_GROUPS)
    total_cols = len(col_order)

    if not alias_map and not rows:
        return

    # ── 문서 제목 행 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=doc_title or "내  역  서")
    _apply_style(title_cell, fill=_FILL_TITLE, font=_FONT_TITLE,
                 align=_ALIGN_CENTER, border=_BORDER_HEADER)
    ws.row_dimensions[1].height = 28

    # ── 그룹 헤더 행 (행2) ──
    col_cursor = 1
    for grp_name, subs in active_groups:
        span = len(subs)
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col_cursor,
                           end_row=2, end_column=col_cursor + span - 1)
        cell = ws.cell(row=2, column=col_cursor, value=grp_name)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        # 병합 오른쪽 셀에도 스타일 적용 (openpyxl 요구사항)
        for extra in range(1, span):
            c = ws.cell(row=2, column=col_cursor + extra)
            _apply_style(c, fill=_FILL_HEADER, font=_FONT_HEADER,
                         align=_ALIGN_CENTER, border=_BORDER_ALL)
        col_cursor += span
    ws.row_dimensions[2].height = 18

    # ── 서브 헤더 행 (행3) ──
    for col_idx, sub_key in enumerate(col_order, start=1):
        # 서브 헤더 표시명: "재료비_단가" → "단가", 단일 컬럼은 그룹명 이미 표시
        label = sub_key.split("_")[1] if "_" in sub_key else ""
        cell = ws.cell(row=3, column=col_idx, value=label)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)

        # 단일 컬럼(label이 빈 경우): 그룹 행(2)과 병합
        if not label:
            ws.merge_cells(start_row=2, start_column=col_idx,
                           end_row=3, end_column=col_idx)
    ws.row_dimensions[3].height = 18

    # ── 데이터 행 ──
    first_key = alias_map.get("품명", headers[0])
    for row_idx, row in enumerate(rows, start=4):
        style = _row_style(row, first_key)
        ws.row_dimensions[row_idx].height = 16

        for col_idx, key in enumerate(col_order, start=1):
            raw_val = _get_detail_row_value(row, key, alias_map)
            cell = ws.cell(row=row_idx, column=col_idx)

            is_money = key.endswith(("_단가", "_금액")) or key == "수량"

            # [수정 D] 금액/단가/수량 컬럼에 숫자 변환 적용
            numeric = _try_parse_number(raw_val) if is_money else None
            if numeric is not None:
                cell.value = numeric
                cell.number_format = '#,##0'
            else:
                cell.value = raw_val

            val = raw_val  # 스타일 판별용

            if style == "section":
                _apply_style(cell, fill=_FILL_SECTION, font=_FONT_SECTION,
                             align=_ALIGN_LEFT, border=_BORDER_ALL)
            elif style == "subtotal":
                _apply_style(cell, fill=_FILL_SUBTOTAL, font=_FONT_SUBTOTAL,
                             align=_ALIGN_RIGHT if (numeric is not None or is_money) else _ALIGN_LEFT,
                             border=_BORDER_ALL)
            else:
                if numeric is not None and is_money:
                    _apply_style(cell, font=_FONT_BODY, align=_ALIGN_RIGHT, border=_BORDER_ALL)
                elif col_idx == 1:
                    _apply_style(cell, font=_FONT_BODY, align=_ALIGN_LEFT, border=_BORDER_ALL)
                else:
                    _apply_style(cell, font=_FONT_BODY, align=_ALIGN_CENTER, border=_BORDER_ALL)

    # ── 컬럼 폭 설정 ──
    _col_widths = {
        "품명": 30, "규격": 14, "단위": 6, "수량": 8,
        "재료비_단가": 11, "재료비_금액": 11,
        "노무비_단가": 11, "노무비_금액": 11,
        "경비_단가": 11,   "경비_금액": 11,
        "합계_단가": 11,   "합계_금액": 11,
        "비고": 12,
    }
    for col_idx, key in enumerate(col_order, start=1):
        width = _col_widths.get(key, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ═══════════════════════════════════════════════════════
# 조건 시트 빌더
# ═══════════════════════════════════════════════════════

def _build_condition_sheet(ws, table: dict):
    """
    조건 시트(일반사항/특기사항) 작성.

    [수정 E] seen_right(전역 set) → prev_row_vals(열 단위 직전 행 비교):
        전역 set은 열/행 구분 없이 누적하여 다른 열·행의 동일 단어까지 삭제했다.
        올바른 suppression은 rowspan 전개로 인한 '같은 열의 수직 반복'만 제거.

    원본: excel_exporter.py L455~489
    """
    headers = table.get("headers", [])
    rows    = table.get("rows", [])

    if not headers:
        return

    # ── 헤더 행 ──
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        ws.column_dimensions[get_column_letter(col_idx)].width = 45
    ws.row_dimensions[1].height = 18

    # ── 데이터 행 ──
    prev_row_vals: dict[int, str] = {}  # 열 인덱스 → 직전 행 값
    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 16
        for col_idx, h in enumerate(headers, start=1):
            val = str(row.get(h, "")).strip()
            # 같은 열의 직전 행과 동일한 값이면 suppression (rowspan 전개 중복 제거)
            if col_idx > 1 and val and val == prev_row_vals.get(col_idx):
                display_val = ""
            else:
                display_val = val
            prev_row_vals[col_idx] = val
            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            _apply_style(cell, font=_FONT_BODY, align=_ALIGN_LEFT, border=_BORDER_ALL)


# ═══════════════════════════════════════════════════════
# 범용 시트 빌더
# ═══════════════════════════════════════════════════════

def _build_generic_sheet(ws, table: dict):
    """
    범용 테이블 시트 — 헤더/데이터를 원본 그대로 기록한다.

    Why: 분류 불가 테이블(BOM, 거래명세서, 공문서 등)을 유실하지 않고
         원데이터를 그대로 Excel에 옮긴다. 사용자가 수동 재편집 가능.

    [수정 C + 수정 H]:
        headers=[], rows=[dict] → rows의 첫 dict 키를 헤더로 자동 생성
        headers=[], rows=[]    → 조용히 스킵

    원본: excel_exporter.py L496~555
    """
    headers = table.get("headers", [])
    rows    = table.get("rows", [])

    # [수정 H] 헤더 없는 경우 폴백
    if not headers and rows:
        if isinstance(rows[0], dict):
            headers = list(rows[0].keys())
        if not headers:
            print(f"    ⚠️ 헤더·키 없는 테이블 스킵: {table.get('table_id', '?')}")
            return
    elif not headers:
        return

    # ── 헤더 행 ──
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
    ws.row_dimensions[1].height = 20

    # ── 데이터 행 ──
    for row_idx, row in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = 16
        for col_idx, h in enumerate(headers, start=1):
            raw_val = str(row.get(h, "")).strip()
            cell = ws.cell(row=row_idx, column=col_idx)

            # [수정 D] 숫자 감지 → 숫자 타입으로 기록 + 콤마 포맷
            numeric = _try_parse_number(raw_val)
            if numeric is not None:
                cell.value = numeric
                cell.number_format = '#,##0'
                _apply_style(cell, font=_FONT_BODY, align=_ALIGN_RIGHT,
                             border=_BORDER_ALL)
            else:
                cell.value = raw_val
                _apply_style(cell, font=_FONT_BODY,
                             align=_ALIGN_CENTER if col_idx == 1 else _ALIGN_LEFT,
                             border=_BORDER_ALL)

    # ── 열 너비 자동 조정 (한글 2바이트 기준) ──
    for col_idx, h in enumerate(headers, start=1):
        header_len = sum(2 if ord(c) > 127 else 1 for c in h)
        width = max(header_len + 4, 10)
        for row in rows:
            val = str(row.get(h, ""))
            val_len = sum(2 if ord(c) > 127 else 1 for c in val)
            width = max(width, val_len + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 50)


# ═══════════════════════════════════════════════════════
# Phase 12.5 신규 빌더 — 본문 / 주석 / 가감산_조건 / 교차참조 / 메타데이터
# ═══════════════════════════════════════════════════════

def _build_text_sheet(ws, sections: list[dict]) -> None:
    """섹션별 clean_text를 행 단위로 덤프."""
    headers = ["섹션 ID", "부문", "장", "제목", "페이지", "본문"]
    widths  = [12, 14, 20, 28, 7, 80]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 18

    for row_idx, s in enumerate(sections, start=2):
        text = s.get("clean_text", "").strip()
        if not text:
            continue
        vals = [
            s.get("section_id", ""), s.get("department", ""),
            s.get("chapter", ""), s.get("title", ""),
            s.get("page", ""), text,
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            align = _ALIGN_LEFT if col_idx == 6 else _ALIGN_CENTER
            _apply_style(cell, font=_FONT_BODY, align=align, border=_BORDER_ALL)


def _build_notes_sheet(ws, sections: list[dict]) -> None:
    """모든 섹션의 notes를 평탄화하여 통합."""
    headers = ["섹션 ID", "섹션 제목", "페이지", "주석 번호", "주석 내용"]
    widths  = [12, 28, 7, 8, 80]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 18

    row_idx = 2
    for s in sections:
        notes = s.get("notes", [])
        if not notes:
            continue
        for note_num, note in enumerate(notes, start=1):
            vals = [s.get("section_id", ""), s.get("title", ""),
                    s.get("page", ""), note_num, note]
            for col_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                align = _ALIGN_LEFT if col_idx == 5 else _ALIGN_CENTER
                _apply_style(cell, font=_FONT_BODY, align=align, border=_BORDER_ALL)
            row_idx += 1


def _build_conditions_sheet(ws, sections: list[dict]) -> None:
    """JSON conditions[]의 {type,condition,rate} dict를 표로 펼침."""
    headers = ["섹션 ID", "섹션 제목", "페이지", "유형", "조건", "비율"]
    widths  = [12, 28, 7, 10, 50, 10]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 18

    row_idx = 2
    for s in sections:
        for cond in s.get("conditions", []):
            vals = [
                s.get("section_id", ""), s.get("title", ""), s.get("page", ""),
                cond.get("type", ""), cond.get("condition", ""), cond.get("rate", ""),
            ]
            for col_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                align = _ALIGN_LEFT if col_idx == 5 else _ALIGN_CENTER
                _apply_style(cell, font=_FONT_BODY, align=align, border=_BORDER_ALL)
            row_idx += 1


def _build_crossref_sheet(ws, sections: list[dict]) -> None:
    """JSON cross_references[]의 {target_section_id,target_chapter,context}를 표로."""
    headers = ["원본 섹션 ID", "원본 섹션 제목", "페이지", "대상 장", "대상 섹션 ID", "참조 문맥"]
    widths  = [12, 28, 7, 14, 14, 60]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 18

    row_idx = 2
    for s in sections:
        for ref in s.get("cross_references", []):
            vals = [
                s.get("section_id", ""), s.get("title", ""), s.get("page", ""),
                ref.get("target_chapter", ""), ref.get("target_section_id", ""),
                ref.get("context", ""),
            ]
            for col_idx, v in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                align = _ALIGN_LEFT if col_idx == 6 else _ALIGN_CENTER
                _apply_style(cell, font=_FONT_BODY, align=align, border=_BORDER_ALL)
            row_idx += 1


def _build_meta_sheet(ws, sections: list[dict]) -> None:
    """섹션별 revision_year / unit_basis를 통합."""
    headers = ["섹션 ID", "섹션 제목", "페이지", "보완연도", "단위 기준"]
    widths  = [12, 28, 7, 10, 14]

    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 18

    row_idx = 2
    for s in sections:
        rev  = s.get("revision_year") or ""
        unit = s.get("unit_basis") or ""
        if not rev and not unit:
            continue
        vals = [s.get("section_id", ""), s.get("title", ""), s.get("page", ""), rev, unit]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            _apply_style(cell, font=_FONT_BODY,
                         align=_ALIGN_CENTER, border=_BORDER_ALL)
        row_idx += 1


# ═══════════════════════════════════════════════════════
# Phase 14 신규 빌더 — 도면 메타데이터
# ═══════════════════════════════════════════════════════

_DRAWING_META_FIELD_ORDER = [
    "dwg_no", "rev", "title", "date", "project", "client", 
    "contractor", "drawn_by", "checked_by", "approved_by", "scale", "sheet"
]

_DRAWING_META_LABELS = {
    "dwg_no": "DWG NO.",
    "rev": "REV.",
    "title": "TITLE",
    "date": "DATE",
    "project": "PROJECT",
    "client": "CLIENT",
    "contractor": "CONTRACTOR",
    "drawn_by": "DRAWN BY",
    "checked_by": "CHECKED BY",
    "approved_by": "APPROVED BY",
    "scale": "SCALE",
    "sheet": "SHEET",
}

def _build_drawing_meta_sheet(ws, drawing_metadata: dict) -> None:
    """
    도면 메타데이터 시트를 수직 형태로 기록한다.
    
    A열: 필드명, B열: 값
    """
    if not drawing_metadata:
        return

    # ── 제목 셀 ──
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1, value="도면 메타데이터")
    _apply_style(title_cell, fill=_FILL_TITLE, font=_FONT_TITLE,
                 align=_ALIGN_CENTER, border=_BORDER_HEADER)
    ws.row_dimensions[1].height = 28

    # ── 헤더 행 ──
    headers = ["필드", "내용"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        _apply_style(cell, fill=_FILL_HEADER, font=_FONT_HEADER,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
    ws.row_dimensions[2].height = 20

    # ── 데이터 행 ──
    row_idx = 3
    for key in _DRAWING_META_FIELD_ORDER:
        val = drawing_metadata.get(key)
        if val is None:
            continue
            
        label = _DRAWING_META_LABELS.get(key, key.upper())
        
        # A열: 필드명
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        _apply_style(cell_a, fill=_FILL_SECTION, font=_FONT_BODY,
                     align=_ALIGN_CENTER, border=_BORDER_ALL)
                     
        # B열: 값
        cell_b = ws.cell(row=row_idx, column=2, value=val)
        _apply_style(cell_b, font=_FONT_BODY,
                     align=_ALIGN_LEFT, border=_BORDER_ALL)
                     
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    # ── 열 너비 설정 ──
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
