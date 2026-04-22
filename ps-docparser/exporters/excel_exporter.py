"""
exporters/excel_exporter.py — Phase 3: JSON → Excel 변환기 (shim + 오케스트레이터)

Why: Phase 12 Step 12-2 분해 결과물.
     스타일 상수/헬퍼 → excel_styles.py
     테이블·행 분류, 숫자 파싱 → excel_classifier.py
     4개 시트 빌더 → excel_builders.py
     이 파일은 기존 import 경로를 100% 유지하는 shim과
     _export_impl() 오케스트레이터 + ExcelExporter 클래스만 포함한다.

공개 API (변경 없음):
    export(sections, output_path, *, title=None) → Path
    ExcelExporter (BaseExporter 구현)

원본: exporters/excel_exporter.py L1~721 (Phase 11 완료 기준)
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

# ── Import shim: 하위 모듈에서 전체 공개/반공개 API re-export ──
from exporters.excel_styles import (
    _THIN, _THICK,
    _BORDER_ALL, _BORDER_HEADER,
    _FILL_HEADER, _FILL_SECTION, _FILL_SUBTOTAL, _FILL_TITLE,
    _FONT_HEADER, _FONT_TITLE, _FONT_SECTION, _FONT_SUBTOTAL,
    _FONT_BODY, _FONT_NOTE,
    _ALIGN_CENTER, _ALIGN_LEFT, _ALIGN_RIGHT,
    _apply_style,
)
from exporters.excel_classifier import (
    _classify_table,
    _is_number,
    _try_parse_number,
    _row_style,
    _RE_NUMBER,
    _RE_NUMERIC,
    _SECTION_KEYWORDS,
    _SUBTOTAL_KEYWORDS,
)
from exporters.excel_builders import (
    _DETAIL_HEADER_GROUPS,
    _build_estimate_sheet,
    _build_detail_sheet,
    _build_condition_sheet,
    _build_generic_sheet,
    _build_text_sheet,
    _build_notes_sheet,
    _build_conditions_sheet,
    _build_crossref_sheet,
    _build_meta_sheet,
    _build_drawing_meta_sheet,
)
from exporters.base_exporter import BaseExporter

__all__ = [
    # excel_styles
    "_THIN", "_THICK", "_BORDER_ALL", "_BORDER_HEADER",
    "_FILL_HEADER", "_FILL_SECTION", "_FILL_SUBTOTAL", "_FILL_TITLE",
    "_FONT_HEADER", "_FONT_TITLE", "_FONT_SECTION", "_FONT_SUBTOTAL",
    "_FONT_BODY", "_FONT_NOTE",
    "_ALIGN_CENTER", "_ALIGN_LEFT", "_ALIGN_RIGHT",
    "_apply_style",
    # excel_classifier
    "_classify_table", "_is_number", "_try_parse_number", "_row_style",
    "_RE_NUMBER", "_RE_NUMERIC", "_SECTION_KEYWORDS", "_SUBTOTAL_KEYWORDS",
    # excel_builders
    "_DETAIL_HEADER_GROUPS",
    "_build_estimate_sheet", "_build_detail_sheet",
    "_build_condition_sheet", "_build_generic_sheet",
    "_build_text_sheet", "_build_notes_sheet", "_build_conditions_sheet",
    "_build_crossref_sheet", "_build_meta_sheet", "_build_drawing_meta_sheet",
    # orchestration (this module)
    "_export_impl", "export", "ExcelExporter",
]


# ═══════════════════════════════════════════════════════
# 오케스트레이터 (이 파일에 유지)
# ═══════════════════════════════════════════════════════

def _export_impl(
    sections: list[dict[str, Any]],
    output_path: str | Path,
    *,
    title: str | None = None,
) -> Path:
    """
    sections JSON을 Excel 파일로 변환한다.

    Args:
        sections:     parse_markdown()이 반환한 섹션 리스트
        output_path:  저장할 .xlsx 경로
        title:        문서 제목 (None이면 section title 또는 파일명 사용)

    Returns:
        저장된 Path 객체

    원본: excel_exporter.py L562~681
    """
    wb = Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    doc_title = title or ""

    # sections 전체를 순회하면서 테이블을 수집·분류
    estimate_tables:  list[dict] = []
    detail_tables:    list[dict] = []
    condition_tables: list[dict] = []
    generic_tables:   list[dict] = []  # [수정 B] generic 대기열 추가

    for section in sections:
        # 문서 제목 우선순위: 인수 > section.title > clean_text 첫 줄
        if not doc_title:
            doc_title = section.get("title", "")
        if not doc_title:
            clean = section.get("clean_text", "")
            doc_title = clean.split("\n")[0].strip()[:60] if clean else ""

        for tbl in section.get("tables", []):
            kind = _classify_table(tbl)
            if kind == "estimate":
                estimate_tables.append(tbl)
            elif kind == "detail":
                detail_tables.append(tbl)
            elif kind == "condition":
                condition_tables.append(tbl)
            elif kind == "generic":          # [수정 B]
                generic_tables.append(tbl)

    # ── 견적서 시트 ──
    if estimate_tables:
        ws_est = wb.create_sheet("견적서")
        ws_est.sheet_view.showGridLines = False
        _build_estimate_sheet(ws_est, estimate_tables[0], doc_title)
        for i, tbl in enumerate(estimate_tables[1:], start=2):
            ws_extra = wb.create_sheet(f"견적서_{i}")
            ws_extra.sheet_view.showGridLines = False
            _build_estimate_sheet(ws_extra, tbl, doc_title)

    # ── 내역서 시트 ──
    if detail_tables:
        ws_det = wb.create_sheet("내역서")
        ws_det.sheet_view.showGridLines = False
        _build_detail_sheet(ws_det, detail_tables[0], doc_title)
        for i, tbl in enumerate(detail_tables[1:], start=2):
            ws_extra = wb.create_sheet(f"내역서_{i}")
            ws_extra.sheet_view.showGridLines = False
            _build_detail_sheet(ws_extra, tbl, doc_title)

    # ── 조건 시트 ──
    if condition_tables:
        ws_cond = wb.create_sheet("조건")
        ws_cond.sheet_view.showGridLines = False
        _build_condition_sheet(ws_cond, condition_tables[0])

    # ── 본문 시트 (Phase 12.5) ──
    if any(s.get("clean_text", "").strip() for s in sections):
        ws_txt = wb.create_sheet("본문")
        ws_txt.sheet_view.showGridLines = False
        _build_text_sheet(ws_txt, sections)

    # ── 주석 시트 ──
    if any(s.get("notes") for s in sections):
        ws_notes = wb.create_sheet("주석")
        ws_notes.sheet_view.showGridLines = False
        _build_notes_sheet(ws_notes, sections)

    # ── 가감산 조건 시트 ──
    if any(s.get("conditions") for s in sections):
        ws_conds = wb.create_sheet("가감산_조건")
        ws_conds.sheet_view.showGridLines = False
        _build_conditions_sheet(ws_conds, sections)

    # ── 교차참조 시트 ──
    if any(s.get("cross_references") for s in sections):
        ws_xref = wb.create_sheet("교차참조")
        ws_xref.sheet_view.showGridLines = False
        _build_crossref_sheet(ws_xref, sections)

    # ── 메타 시트 ──
    if any((s.get("revision_year") or s.get("unit_basis")) for s in sections):
        ws_meta = wb.create_sheet("메타데이터")
        ws_meta.sheet_view.showGridLines = False
        _build_meta_sheet(ws_meta, sections)

    # ── 도면_메타 시트 (Phase 14) ──
    drawing_meta_sections = [s for s in sections if s.get("type") == "drawing_meta"]
    if drawing_meta_sections:
        ws_dwg_meta = wb.create_sheet("도면_메타")
        ws_dwg_meta.sheet_view.showGridLines = False
        _build_drawing_meta_sheet(ws_dwg_meta, drawing_meta_sections[0].get("drawing_metadata", {}))

    # ── 범용 시트 (분류 불가 테이블) [수정 B] ──
    if generic_tables:
        for i, tbl in enumerate(generic_tables, start=1):
            tbl_title = tbl.get("title", "")
            if tbl_title:
                sheet_name = (
                    f"{tbl_title}_{i}"
                    if len(generic_tables) > 1
                    and sum(1 for t in generic_tables if t.get("title") == tbl_title) > 1
                    else tbl_title
                )
            else:
                sheet_name = f"Table_{i}" if len(generic_tables) > 1 else "Table"

            ws_gen = wb.create_sheet(sheet_name[:31])  # Excel 시트명 31자 제한
            ws_gen.sheet_view.showGridLines = False
            _build_generic_sheet(ws_gen, tbl)

    # ── 분류된 테이블/필드가 하나도 없을 때 ── 원시 덤프
    has_any_legacy = estimate_tables or detail_tables or condition_tables or generic_tables
    has_any_new = (
        any(s.get("clean_text", "").strip() for s in sections)
        or any(s.get("notes") for s in sections)
        or any(s.get("conditions") for s in sections)
        or any(s.get("cross_references") for s in sections)
        or any((s.get("revision_year") or s.get("unit_basis")) for s in sections)
        or bool(drawing_meta_sections)
    )
    if not has_any_legacy and not has_any_new:
        ws_raw = wb.create_sheet("데이터")
        ws_raw.cell(row=1, column=1,
                    value="⚠ 분류 가능한 데이터가 없습니다.")
        ws_raw.cell(row=1, column=1).font = Font(color="FF0000", bold=True)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # [수정 G] wb.save() PermissionError 처리
    try:
        wb.save(output_path)
    except PermissionError as e:
        from utils.io import ParserError
        raise ParserError(
            f"파일 저장 (권한 거부): {output_path.name}\n"
            f"  → 해당 파일이 Excel 등 다른 프로그램에서 열려 있는지 확인하세요.\n"
            f"  상세: {e}"
        )
    except OSError as e:
        from utils.io import ParserError
        raise ParserError(
            f"파일 저장 (I/O 오류): {output_path.name}\n"
            f"  → 디스크 공간이나 경로 길이를 확인하세요.\n"
            f"  상세: {e}"
        )

    return output_path


# ── 하위 호환 공개 API 별칭 ──
# Why: 기존 코드(main.py, _test_phase3.py)에서
#      from exporters.excel_exporter import export as excel_export 사용.
export = _export_impl


# ═══════════════════════════════════════════════════════
# BaseExporter 클래스 인터페이스
# ═══════════════════════════════════════════════════════

class ExcelExporter(BaseExporter):
    """JSON 섹션 리스트를 Excel 워크북으로 변환한다."""

    file_extension = ".xlsx"

    def export(
        self,
        sections: list[dict],
        output_path: Path,
        *,
        metadata: dict | None = None,
        preset_config: dict | None = None,
    ) -> Path:
        """
        BaseExporter 인터페이스 구현.

        현재는 _export_impl()에 위임한다.
        preset_config가 있으면 향후 _write_preset_sheets()로 분기.

        Why: _export_impl()은 수정 A~H로 실전 검증되었다.
             전면 리팩터 대신 위임 패턴으로 안전하게 클래스 인터페이스를 제공.
             Phase 4 이후 점진적으로 로직을 클래스 내부로 이전한다.
        """
        title = metadata.get("description") if metadata else None
        return _export_impl(sections, output_path, title=title)
