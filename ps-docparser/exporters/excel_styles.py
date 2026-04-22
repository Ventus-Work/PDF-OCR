"""
exporters/excel_styles.py — Excel 셀 스타일 상수 및 적용 헬퍼

Why: Phase 12 Step 12-2 분해 결과물.
     excel_exporter.py의 스타일 상수(Fill/Font/Alignment/Border)와
     _apply_style() 헬퍼를 분리한 순수 스타일 모듈.
     openpyxl 스타일 객체 외 어떤 비즈니스 로직도 포함하지 않는다.

원본: exporters/excel_exporter.py L53~72 (상수) + L204~212 (_apply_style)
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ═══════════════════════════════════════════════════════
# 테두리 상수
# ═══════════════════════════════════════════════════════

_THIN  = Side(style="thin")
_THICK = Side(style="medium")
_BORDER_ALL    = Border(left=_THIN,  right=_THIN,  top=_THIN,  bottom=_THIN)
_BORDER_HEADER = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)

# ═══════════════════════════════════════════════════════
# Fill 상수
# ═══════════════════════════════════════════════════════

_FILL_HEADER   = PatternFill("solid", fgColor="1F3864")  # 진남색
_FILL_SECTION  = PatternFill("solid", fgColor="D6E4F0")  # 연청색 (구분행)
_FILL_SUBTOTAL = PatternFill("solid", fgColor="FFF2CC")  # 연노랑 (소계/합계행)
_FILL_TITLE    = PatternFill("solid", fgColor="2E75B6")  # 중간 파랑 (문서 제목)

# ═══════════════════════════════════════════════════════
# Font 상수
# ═══════════════════════════════════════════════════════

_FONT_HEADER   = Font(name="맑은 고딕", bold=True,  color="FFFFFF", size=9)
_FONT_TITLE    = Font(name="맑은 고딕", bold=True,  color="FFFFFF", size=11)
_FONT_SECTION  = Font(name="맑은 고딕", bold=True,  size=9)
_FONT_SUBTOTAL = Font(name="맑은 고딕", bold=True,  color="C00000", size=9)
_FONT_BODY     = Font(name="맑은 고딕", size=9)
_FONT_NOTE     = Font(name="맑은 고딕", italic=True, size=8, color="595959")

# ═══════════════════════════════════════════════════════
# Alignment 상수
# ═══════════════════════════════════════════════════════

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")


# ═══════════════════════════════════════════════════════
# 스타일 적용 헬퍼
# ═══════════════════════════════════════════════════════

def _apply_style(cell, fill=None, font=None, align=None, border=_BORDER_ALL):
    """
    단일 셀에 fill/font/alignment/border를 일괄 적용한다.

    Why: 4개 속성을 개별로 설정하면 호출부가 장황해진다.
         None이면 해당 속성을 건드리지 않아 기존 스타일을 보존한다.

    원본: excel_exporter.py L204~212
    """
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
