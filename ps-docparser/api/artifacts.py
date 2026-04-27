"""Artifact and QA readers for UI job results."""

from __future__ import annotations

import base64
import json
import re
from pathlib import Path
from typing import Any

from .errors import ApiError
from .jobs import JobRecord
from .schemas import ArtifactItem, ArtifactPreviewResponse, ArtifactsResponse, QAResponse

ALLOWED_DOWNLOAD_SUFFIXES = {".md", ".json", ".xlsx"}
DENIED_DOWNLOAD_SUFFIXES = {".env", ".py", ".db", ".log"}
SPECIAL_KINDS = {
    "RUN_MANIFEST.json": "manifest",
    "RUN_SUMMARY.md": "summary",
    "OUTPUT_QA_REPORT.md": "qa",
}


def _safe_read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig", errors="replace")


def encode_artifact_id(relative_path: str) -> str:
    raw = relative_path.replace("\\", "/").encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def decode_artifact_id(artifact_id: str) -> str:
    padding = "=" * (-len(artifact_id) % 4)
    try:
        value = base64.urlsafe_b64decode((artifact_id + padding).encode("ascii")).decode("utf-8")
    except Exception as exc:
        raise ApiError(status_code=404, code="artifact_not_found", message="산출물을 찾을 수 없습니다.") from exc
    return value.replace("\\", "/")


def ensure_inside(path: Path, base_dir: Path) -> Path:
    resolved = path.resolve()
    base = base_dir.resolve()
    try:
        resolved.relative_to(base)
    except ValueError as exc:
        raise ApiError(status_code=400, code="unsafe_path", message="허용되지 않는 산출물 경로입니다.") from exc
    return resolved


def artifact_kind(path: Path) -> str:
    if path.name in SPECIAL_KINDS:
        return SPECIAL_KINDS[path.name]
    suffix = path.suffix.lower().lstrip(".")
    if suffix in {"md", "json", "xlsx"}:
        return suffix
    return "other"


def _load_manifest_metadata(result_dir: Path) -> dict[str, dict[str, str]]:
    manifest_path = result_dir / "RUN_MANIFEST.json"
    if not manifest_path.exists():
        return {}
    try:
        manifest = json.loads(_safe_read_text(manifest_path))
    except Exception:
        return {}

    metadata: dict[str, dict[str, str]] = {}

    def collect(artifact: dict[str, Any]) -> None:
        role = str(artifact.get("role") or "unknown")
        domain = str(artifact.get("domain") or "unknown")
        quality = str(artifact.get("quality_status") or "unknown")
        for key in ("md", "json", "xlsx"):
            rel = artifact.get(key)
            if rel:
                metadata[str(rel).replace("\\", "/")] = {
                    "role": role,
                    "domain": domain,
                    "quality_status": quality,
                }

    for entry in manifest.get("inputs", []) or []:
        primary = entry.get("primary")
        if isinstance(primary, dict):
            collect(primary)
        for diagnostic in entry.get("diagnostics", []) or []:
            if isinstance(diagnostic, dict):
                collect(diagnostic)
    return metadata


def _fallback_metadata_for_path(relative_path: str) -> dict[str, str]:
    normalized = relative_path.replace("\\", "/")
    if normalized.startswith("_compare/"):
        domain = "generic"
        if "/route_manifest.json" in normalized:
            domain = "route"
        elif "/bom/" in normalized or normalized.endswith("_bom.json"):
            domain = "bom"
        elif "/estimate/" in normalized or "estimate" in normalized:
            domain = "estimate"
        return {
            "role": "compare",
            "domain": domain,
            "quality_status": "unknown",
        }
    return {}


def list_artifacts(record: JobRecord) -> ArtifactsResponse:
    result_dir = record.result_dir
    if not result_dir.exists():
        return ArtifactsResponse(
            job_id=record.job_id,
            artifacts=[],
            message="결과 폴더가 아직 생성되지 않았습니다.",
        )

    manifest_metadata = _load_manifest_metadata(result_dir)
    files = [
        path
        for path in result_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in ALLOWED_DOWNLOAD_SUFFIXES
    ]

    def sort_key(path: Path) -> tuple[int, str]:
        special_rank = {"RUN_MANIFEST.json": 0, "RUN_SUMMARY.md": 1, "OUTPUT_QA_REPORT.md": 2}
        return (special_rank.get(path.name, 10), str(path.relative_to(result_dir)).lower())

    artifacts: list[ArtifactItem] = []
    for path in sorted(files, key=sort_key):
        rel = str(path.relative_to(result_dir)).replace("\\", "/")
        meta = manifest_metadata.get(rel, {}) or _fallback_metadata_for_path(rel)
        artifacts.append(
            ArtifactItem(
                artifact_id=encode_artifact_id(rel),
                name=path.name,
                relative_path=rel,
                kind=artifact_kind(path),  # type: ignore[arg-type]
                size_bytes=path.stat().st_size,
                download_url=f"/api/jobs/{record.job_id}/artifacts/{encode_artifact_id(rel)}",
                role=meta.get("role", "unknown"),
                domain=meta.get("domain", "unknown"),
                quality_status=meta.get("quality_status", "unknown"),
            )
        )
    return ArtifactsResponse(job_id=record.job_id, artifacts=artifacts)


def resolve_artifact_path(record: JobRecord, artifact_id: str) -> Path:
    rel = decode_artifact_id(artifact_id)
    if rel.startswith("/") or re.match(r"^[A-Za-z]:", rel) or ".." in Path(rel).parts:
        raise ApiError(status_code=400, code="unsafe_path", message="허용되지 않는 산출물 경로입니다.")
    path = ensure_inside(record.result_dir / rel, record.result_dir)
    if not path.exists() or not path.is_file():
        raise ApiError(status_code=404, code="artifact_not_found", message="산출물을 찾을 수 없습니다.")
    suffix = path.suffix.lower()
    if suffix in DENIED_DOWNLOAD_SUFFIXES or suffix not in ALLOWED_DOWNLOAD_SUFFIXES:
        raise ApiError(status_code=404, code="artifact_not_found", message="다운로드할 수 없는 산출물입니다.")
    return path


def preview_artifact(record: JobRecord, artifact_id: str) -> ArtifactPreviewResponse:
    path = resolve_artifact_path(record, artifact_id)
    kind = artifact_kind(path)
    if kind in {"json", "manifest"}:
        text = _safe_read_text(path)
        try:
            data = json.loads(text)
            pretty = json.dumps(data, ensure_ascii=False, indent=2)
        except Exception:
            data = None
            pretty = text
        return ArtifactPreviewResponse(
            job_id=record.job_id,
            artifact_id=artifact_id,
            name=path.name,
            kind=kind,  # type: ignore[arg-type]
            text=pretty,
            json_data=data,
        )
    if kind in {"md", "summary", "qa"}:
        return ArtifactPreviewResponse(
            job_id=record.job_id,
            artifact_id=artifact_id,
            name=path.name,
            kind=kind,  # type: ignore[arg-type]
            text=_safe_read_text(path),
        )
    if kind == "xlsx":
        return _preview_xlsx(record, artifact_id, path)
    raise ApiError(status_code=404, code="preview_not_supported", message="미리보기를 지원하지 않는 산출물입니다.")


def _preview_xlsx(record: JobRecord, artifact_id: str, path: Path) -> ArtifactPreviewResponse:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(status_code=500, code="preview_dependency_missing", message="Excel 미리보기에 openpyxl이 필요합니다.") from exc

    max_rows = 100
    max_columns = 50
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        first_row = next(iterator, None)
        columns = [
            str(value) if value not in (None, "") else f"Column {index + 1}"
            for index, value in enumerate((first_row or [])[:max_columns])
        ]
        rows: list[list[Any]] = []
        truncated = False
        for row in iterator:
            if len(rows) >= max_rows:
                truncated = True
                break
            rows.append([_cell_value(value) for value in row[:max_columns]])
        return ArtifactPreviewResponse(
            job_id=record.job_id,
            artifact_id=artifact_id,
            name=path.name,
            kind="xlsx",
            columns=columns,
            rows=rows,
            truncated=truncated,
            message=f"{sheet.title} 시트의 상위 {max_rows}행까지 표시합니다.",
        )
    finally:
        workbook.close()


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def parse_qa(record: JobRecord) -> QAResponse:
    result_dir = record.result_dir
    report_path = result_dir / "OUTPUT_QA_REPORT.md"
    manifest_path = result_dir / "RUN_MANIFEST.json"

    summary = QAResponse(job_id=record.job_id, status="unknown")

    if manifest_path.exists():
        manifest_summary = _manifest_counts(manifest_path)
        summary.has_manifest = True
        summary.manifest_inputs = manifest_summary["inputs"]
        summary.manifest_representative = manifest_summary["representative"]
        summary.manifest_diagnostic = manifest_summary["diagnostic"]
        summary.manifest_domains = manifest_summary["domains"]

    if not report_path.exists():
        return summary

    text = _safe_read_text(report_path)
    summary.summary_markdown = text
    summary.report_path = str(report_path.relative_to(result_dir))

    bullet_values = _parse_report_bullets(text)
    raw_status = bullet_values.get("Status", "unknown").lower()
    summary.status = {"pass": "ok", "ok": "ok", "warn": "warn", "fail": "fail"}.get(raw_status, "unknown")  # type: ignore[assignment]
    summary.json_files = _to_int(bullet_values.get("JSON files"))
    summary.excel_files = _to_int(bullet_values.get("Excel files"))
    summary.has_manifest = _to_bool(bullet_values.get("RUN_MANIFEST.json")) or summary.has_manifest
    summary.manifest_inputs = _to_int(bullet_values.get("Manifest inputs")) or summary.manifest_inputs
    summary.manifest_representative = _to_int(bullet_values.get("Manifest representative")) or summary.manifest_representative
    summary.manifest_diagnostic = _to_int(bullet_values.get("Manifest diagnostic")) or summary.manifest_diagnostic
    summary.header_key_mismatch = _to_int(bullet_values.get("Header/key mismatch"))
    summary.bad_composite_headers = _to_int(bullet_values.get("Bad composite headers"))

    quality_warnings = _parse_section_counts(text, "## Quality Warnings")
    manifest_domains = _parse_section_counts(text, "## Manifest Domains")
    if quality_warnings:
        summary.quality_warnings = quality_warnings
    if manifest_domains:
        summary.manifest_domains = manifest_domains
    return summary


def _manifest_counts(manifest_path: Path) -> dict[str, Any]:
    try:
        manifest = json.loads(_safe_read_text(manifest_path))
    except Exception:
        return {"inputs": 0, "representative": 0, "diagnostic": 0, "domains": {}}

    representative = 0
    diagnostic = 0
    domains: dict[str, int] = {}
    inputs = manifest.get("inputs", []) or []
    for entry in inputs:
        primary = entry.get("primary") or {}
        if primary.get("role") == "representative":
            representative += 1
            domain = str(primary.get("domain") or "unknown")
            domains[domain] = domains.get(domain, 0) + 1
        for diagnostic_artifact in entry.get("diagnostics", []) or []:
            diagnostic += 1
            if diagnostic_artifact.get("role") == "representative":
                representative += 1
                domain = str(diagnostic_artifact.get("domain") or "unknown")
                domains[domain] = domains.get(domain, 0) + 1
    return {
        "inputs": len(inputs),
        "representative": representative,
        "diagnostic": diagnostic,
        "domains": domains,
    }


def _parse_report_bullets(text: str) -> dict[str, str]:
    values: dict[str, str] = {}
    for line in text.splitlines():
        match = re.match(r"-\s+([^:]+):\s+`?([^`]+)`?", line.strip())
        if match:
            values[match.group(1).strip()] = match.group(2).strip()
    return values


def _parse_section_counts(text: str, heading: str) -> dict[str, int]:
    in_section = False
    counts: dict[str, int] = {}
    for line in text.splitlines():
        stripped = line.strip()
        if stripped == heading:
            in_section = True
            continue
        if in_section and stripped.startswith("## "):
            break
        if not in_section or not stripped.startswith("- "):
            continue
        if stripped == "- none":
            continue
        match = re.match(r"-\s+`?([^`:]+)`?:\s+(\d+)", stripped)
        if match:
            counts[match.group(1).strip()] = int(match.group(2))
    return counts


def _to_int(value: str | None) -> int:
    if value is None:
        return 0
    match = re.search(r"\d+", value)
    return int(match.group(0)) if match else 0


def _to_bool(value: str | None) -> bool:
    return str(value or "").strip().lower() in {"yes", "true", "1", "y"}
