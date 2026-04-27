from __future__ import annotations

import json
import sys
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock

import openpyxl

import main as cli_main


def _load_real_bom_sections(fixtures_dir: Path) -> list[dict]:
    fixture_path = fixtures_dir / "mock_responses" / "bom_real_sample_sections.json"
    return json.loads(fixture_path.read_text(encoding="utf-8"))


def _make_bom_result() -> SimpleNamespace:
    return SimpleNamespace(
        raw_text=("quote item qty unit price amount memo " * 40).strip(),
        bom_sections=[],
        line_list_sections=[],
        drawing_metadata={"dwg_no": None},
    )


def test_main_cli_bom_preset_material_quote_creates_fallback_artifacts(
    sample_pdf_dir: Path,
    fixtures_dir: Path,
    tmp_path: Path,
    mocker,
):
    input_pdf = sample_pdf_dir / "minimal.pdf"
    output_dir = tmp_path / "output"
    sections = _load_real_bom_sections(fixtures_dir)

    mock_engine = MagicMock()
    mock_engine.supports_ocr = True

    mocker.patch.object(cli_main, "validate_config", return_value={"errors": []})
    mocker.patch.object(cli_main, "_init_cache", return_value=None)
    mocker.patch.object(cli_main, "_project_root", tmp_path)
    mocker.patch("pipelines.bom_pipeline.create_engine", return_value=mock_engine)
    mocker.patch("presets.bom.get_bom_keywords", return_value={})
    mocker.patch("presets.bom.get_image_settings", return_value={})
    mocker.patch(
        "extractors.bom_extractor.extract_bom_with_retry",
        return_value=_make_bom_result(),
    )
    mocker.patch("extractors.bom_extractor.to_sections", return_value=sections)
    mocker.patch("pipelines.bom_pipeline.detect_material_quote", return_value=True)
    process_doc = mocker.patch(
        "extractors.ocr_document_extractor.process_pdf_ocr_document",
        return_value="# fallback estimate",
    )
    mocker.patch("parsers.document_parser.parse_markdown", return_value=sections)
    mocker.patch("presets.estimate.extract_cover_metadata", return_value={})
    mocker.patch.object(
        sys,
        "argv",
        [
            "main.py",
            str(input_pdf),
            "--preset",
            "bom",
            "--engine",
            "zai",
            "--output",
            "excel",
            "--output-dir",
            str(output_dir),
        ],
    )

    cli_main.main()

    assert process_doc.call_count == 1
    assert (tmp_path / "ps-docparser.log").exists()

    main_md_files = list(output_dir.glob("*_minimal_bom.md"))
    main_json_files = list(output_dir.glob("*_minimal_bom.json"))
    main_xlsx_files = list(output_dir.glob("*_minimal_bom.xlsx"))
    fallback_md_files = list(output_dir.glob("*_minimal_bom_fallback_estimate.md"))
    fallback_json_files = list(output_dir.glob("*_minimal_bom_fallback_estimate.json"))
    fallback_xlsx_files = list(output_dir.glob("*_minimal_bom_fallback_estimate.xlsx"))

    assert len(main_md_files) == 1
    assert len(main_json_files) == 1
    assert len(main_xlsx_files) == 1
    assert len(fallback_md_files) == 1
    assert len(fallback_json_files) == 1
    assert len(fallback_xlsx_files) == 1
    assert not (output_dir / "_compare").exists()

    main_sections = json.loads(main_json_files[0].read_text(encoding="utf-8-sig"))
    fallback_sections = json.loads(fallback_json_files[0].read_text(encoding="utf-8-sig"))
    assert main_sections == sections
    assert fallback_sections == sections

    main_wb = openpyxl.load_workbook(main_xlsx_files[0])
    fallback_wb = openpyxl.load_workbook(fallback_xlsx_files[0])
    assert main_wb.sheetnames == ["BOM_자재표"]
    assert fallback_wb.sheetnames == ["BOM_자재표"]


def test_main_cli_bom_preset_no_bom_fallback_skips_estimate_rerun(
    sample_pdf_dir: Path,
    fixtures_dir: Path,
    tmp_path: Path,
    mocker,
):
    input_pdf = sample_pdf_dir / "minimal.pdf"
    output_dir = tmp_path / "output"
    sections = _load_real_bom_sections(fixtures_dir)

    mock_engine = MagicMock()
    mock_engine.supports_ocr = True

    mocker.patch.object(cli_main, "validate_config", return_value={"errors": []})
    mocker.patch.object(cli_main, "_init_cache", return_value=None)
    mocker.patch.object(cli_main, "_project_root", tmp_path)
    mocker.patch("pipelines.bom_pipeline.create_engine", return_value=mock_engine)
    mocker.patch("presets.bom.get_bom_keywords", return_value={})
    mocker.patch("presets.bom.get_image_settings", return_value={})
    mocker.patch(
        "extractors.bom_extractor.extract_bom_with_retry",
        return_value=_make_bom_result(),
    )
    mocker.patch("extractors.bom_extractor.to_sections", return_value=sections)
    mocker.patch("pipelines.bom_pipeline.detect_material_quote", return_value=True)
    process_doc = mocker.patch(
        "extractors.ocr_document_extractor.process_pdf_ocr_document",
        return_value="# fallback estimate",
    )
    mocker.patch.object(
        sys,
        "argv",
        [
            "main.py",
            str(input_pdf),
            "--preset",
            "bom",
            "--engine",
            "zai",
            "--output",
            "excel",
            "--output-dir",
            str(output_dir),
            "--no-bom-fallback",
        ],
    )

    cli_main.main()

    assert process_doc.call_count == 0
    assert (tmp_path / "ps-docparser.log").exists()

    main_md_files = list(output_dir.glob("*_minimal_bom.md"))
    main_json_files = list(output_dir.glob("*_minimal_bom.json"))
    main_xlsx_files = list(output_dir.glob("*_minimal_bom.xlsx"))
    fallback_md_files = list(output_dir.glob("*_minimal_bom_fallback_estimate.md"))
    fallback_json_files = list(output_dir.glob("*_minimal_bom_fallback_estimate.json"))
    fallback_xlsx_files = list(output_dir.glob("*_minimal_bom_fallback_estimate.xlsx"))

    assert len(main_md_files) == 1
    assert len(main_json_files) == 1
    assert len(main_xlsx_files) == 1
    assert not fallback_md_files
    assert not fallback_json_files
    assert not fallback_xlsx_files
    assert not (output_dir / "_compare").exists()

    main_sections = json.loads(main_json_files[0].read_text(encoding="utf-8-sig"))
    assert main_sections == sections

    main_wb = openpyxl.load_workbook(main_xlsx_files[0])
    assert main_wb.sheetnames == ["BOM_자재표"]
