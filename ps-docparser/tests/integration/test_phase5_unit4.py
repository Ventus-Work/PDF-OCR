"""
test_phase5_unit4.py — 단위 4 (Excel 집계 출력) 검증

실행: python test_phase5_unit4.py
의존성: openpyxl (pip install openpyxl)
"""

import sys
import json
import tempfile
from pathlib import Path

_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

results = []
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


def check(name: str, ok: bool, detail: str = ""):
    status = "PASS" if ok else "FAIL"
    results.append((name, status, detail))
    mark = "[OK]" if ok else "[XX]"
    print(f"  {mark} {name}" + (f"  --  {detail}" if detail else ""))


def run_all():
    print("\n" + "=" * 62)
    print("  단위 4: BOM 집계 Excel 출력 연동 검증")
    print("=" * 62 + "\n")

    # ────────────────────────────────────────────────────────
    # TC-1: aggregate_boms() 반환값에 headers/rows 포맷 병용여부
    # ────────────────────────────────────────────────────────
    print("[TC-1] aggregate_boms() 반환 테이블 포맷 (array + headers/rows 병용)")
    from exporters.bom_aggregator import aggregate_boms

    doc = [{
        "type": "estimate",
        "tables": [{
            "array": [
                ["ITEM_NO", "DESCRIPTION", "SIZE", "MAT'L", "Q'TY", "WT(KG)"],
                ["1", "PIPE CLAMP", "50A", "SS400", "4", "2.0"],
                ["2", "BASE PLATE", "100x100", "SUS304", "1", "1.5"],
            ]
        }]
    }]

    with tempfile.TemporaryDirectory() as tmpdir:
        p = Path(tmpdir) / "test.json"
        with open(p, "w", encoding="utf-8") as f:
            json.dump(doc, f)

        sections = aggregate_boms([p])
        table = sections[0]["tables"][0]

        check("반환 테이블에 'headers' 키 존재", "headers" in table)
        check("반환 테이블에 'rows' 키 존재", "rows" in table)
        check("반환 테이블에 'array' 키 존재 (하위 호환)", "array" in table)

        headers = table.get("headers", [])
        rows = table.get("rows", [])
        check("headers가 리스트 타입", isinstance(headers, list))
        check("rows가 리스트 타입", isinstance(rows, list))
        check("rows가 dict 리스트", all(isinstance(r, dict) for r in rows))

        check("ITEM_NO 헤더 포함", "ITEM_NO" in headers)
        check("Q'TY 헤더 포함", "Q'TY" in headers)
        check("WT(KG) 헤더 포함", "WT(KG)" in headers)

        # rows dict 키가 headers와 일치하는지 확인
        if rows:
            row_keys = set(rows[0].keys())
            header_set = set(headers)
            check("rows[0] 키가 headers와 일치", row_keys == header_set,
                  f"row_keys={row_keys}, headers={header_set}")

        # ITEM_NO 일련번호 부여 확인
        if rows and len(rows) >= 2:
            check("ITEM_NO 일련번호 1부터 시작", rows[0].get("ITEM_NO") == 1)
            check("ITEM_NO 순차 증가", rows[1].get("ITEM_NO") == 2)

    # ────────────────────────────────────────────────────────
    # TC-2: export_aggregated_excel() 함수 존재 및 시그니처
    # ────────────────────────────────────────────────────────
    print("\n[TC-2] export_aggregated_excel() 함수 시그니처")
    import inspect
    from exporters.bom_aggregator import export_aggregated_excel

    check("export_aggregated_excel 함수 존재", callable(export_aggregated_excel))
    sig = inspect.signature(export_aggregated_excel)
    params = list(sig.parameters.keys())
    check("파라미터: json_files", "json_files" in params)
    check("파라미터: output_path", "output_path" in params)
    check("파라미터: title (선택)", "title" in params)

    # ────────────────────────────────────────────────────────
    # TC-3: export_aggregated_excel() 실제 xlsx 생성 확인
    # ────────────────────────────────────────────────────────
    print("\n[TC-3] export_aggregated_excel() → xlsx 생성 실전 테스트")

    doc_a = [{
        "type": "bom",
        "tables": [{
            "array": [
                ["ITEM NO.", "DESCRIPTION", "SIZE", "MATERIAL", "Q'TY", "WT(kg)"],
                ["1", "U-BOLT", "40A", "SS400", "8", "0.8"],
                ["2", "SADDLE", "40A", "SS400", "4", "1.2"],
            ]
        }]
    }]
    doc_b = [{
        "type": "estimate",
        "tables": [{
            "array": [
                ["NO", "DESC", "SIZE", "MAT'L", "Qty", "WT(KG)"],
                ["1", "U-BOLT", "40A", "SS400", "4", "0.8"],   # 도면A와 병합 → Q'TY:12, WT:1.6
                ["2", "CLAMP", "80A", "SUS304", "2", "3.5"],   # 신규
            ]
        }]
    }]

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_dir = Path(tmpdir)
        pA = tmp_dir / "drawing_A_bom.json"
        pB = tmp_dir / "drawing_B_bom.json"
        with open(pA, "w", encoding="utf-8") as f:
            json.dump(doc_a, f)
        with open(pB, "w", encoding="utf-8") as f:
            json.dump(doc_b, f)

        xlsx_path = tmp_dir / "BOM집계.xlsx"

        try:
            result = export_aggregated_excel([pA, pB], xlsx_path, title="테스트 집계")
            check("xlsx 파일 생성 성공", result.exists(), str(result))
            check("반환값이 Path 인스턴스", isinstance(result, Path))
            check("xlsx 확장자 확인", result.suffix == ".xlsx")

            # openpyxl로 내용 검증
            try:
                import openpyxl
                wb = openpyxl.load_workbook(xlsx_path)
                sheets = wb.sheetnames
                check("워크북에 시트 1개 이상 존재", len(sheets) >= 1, f"시트: {sheets}")

                # 집계 결과 행 수 확인:
                # doc_a: U-BOLT(40A/SS400), SADDLE(40A/SS400) → (40A, SS400) 1그룹
                # doc_b: U-BOLT(40A/SS400) → 위 그룹에 합산, CLAMP(80A/SUS304) → 신규
                # 결과: (40A, SS400) + (80A, SUS304) = 2개 항목
                # 워크시트: 헤더 1행 + 데이터 2행 = 3행 (이것이 정상)
                ws = wb.active
                data_rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
                check("데이터 행이 3행 이상 (헤더+2항목)", len(data_rows) >= 3,
                      f"실제 행 수: {len(data_rows)}")
                wb.close()

            except ImportError:
                check("openpyxl 설치 확인", False, "pip install openpyxl 필요")

        except Exception as e:
            check("export_aggregated_excel() 실행", False, str(e))
            import traceback
            traceback.print_exc()

    # ────────────────────────────────────────────────────────
    # TC-4: main.py에 배치 집계 hook 코드 존재
    # ────────────────────────────────────────────────────────
    print("\n[TC-4] main.py 배치 집계 hook 코드 검사")
    main_src = (_HERE / "main.py").read_text(encoding="utf-8")
    check(
        "export_aggregated_excel 호출 코드 존재",
        "export_aggregated_excel" in main_src,
    )
    check(
        "preset == 'bom' 조건부 집계 분기 존재",
        "args.preset == \"bom\"" in main_src,
    )
    check(
        "globals('*.json') 역추적 코드 존재",
        "_bom.json\")" in main_src,
    )
    check(
        "BOM집계.xlsx 출력 파일명 코드 존재",
        "BOM집계" in main_src,
    )

    # ── 최종 결과 ──
    _print_summary()


def _print_summary():
    total = len(results)
    passed = sum(1 for _, s, _ in results if s == "PASS")
    failed = total - passed

    print("\n" + "=" * 62)
    print(f"  결과: {passed}/{total} 통과 / {failed}건 실패")
    print("=" * 62)

    if failed:
        print("\n[실패 항목]")
        for name, status, detail in results:
            if status == "FAIL":
                print(f"  - {name}" + (f": {detail}" if detail else ""))
        sys.exit(1)
    else:
        print("\n[완료] 단위 4 검증 -- 모든 테스트 통과")
        sys.exit(0)


if __name__ == "__main__":
    run_all()
