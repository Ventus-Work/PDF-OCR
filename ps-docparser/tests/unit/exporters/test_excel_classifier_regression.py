from pathlib import Path

import openpyxl

from exporters.excel_exporter import ExcelExporter, _classify_table


def test_estimate_detail_condition_document_layout(tmp_path: Path):
    sections = [
        {
            "title": "doc",
            "clean_text": "body",
            "tables": [
                {
                    "headers": [
                        "\u540d\u7a31",
                        "\u898f\u683c",
                        "\u55ae\u4f4d",
                        "\u6578\u91cf",
                        "\u55ae\u50f9",
                        "\u91d1\u984d",
                        "\u5099\u8003",
                    ],
                    "rows": [
                        {
                            "\u540d\u7a31": "PIPE",
                            "\u898f\u683c": "150A",
                            "\u55ae\u4f4d": "EA",
                            "\u6578\u91cf": 4,
                            "\u55ae\u50f9": 1000,
                            "\u91d1\u984d": 4000,
                            "\u5099\u8003": "",
                        }
                    ],
                    "title": "estimate",
                },
                {
                    "headers": ["\ud488\uba85", "\uaddc\uaca9", "\ud569\uacc4_\uae08\uc561"],
                    "rows": [
                        {"\ud488\uba85": "PIPE", "\uaddc\uaca9": "150A", "\ud569\uacc4_\uae08\uc561": 4000}
                    ],
                    "title": "detail",
                },
                {
                    "headers": ["\uc77c\ubc18\uc0ac\ud56d", "\ud2b9\uae30\uc0ac\ud56d"],
                    "rows": [
                        {
                            "\uc77c\ubc18\uc0ac\ud56d": "SS400",
                            "\ud2b9\uae30\uc0ac\ud56d": "\uc6a9\uc811",
                        }
                    ],
                    "type": "D_\uae30\ud0c0",
                    "title": "condition",
                },
            ],
        }
    ]
    out = tmp_path / "estimate_detail_condition.xlsx"
    ExcelExporter().export(sections, out)

    wb = openpyxl.load_workbook(out)
    assert "\uacac\uc801\uc11c" in wb.sheetnames
    assert "\ub0b4\uc5ed\uc11c" in wb.sheetnames
    assert "\uc870\uac74" in wb.sheetnames
    assert "\ubcf8\ubb38" in wb.sheetnames


def test_classify_hanja_summary_table_as_estimate():
    table = {
        "headers": [
            "\u540d\u7a31",
            "\u898f\u683c",
            "\u55ae\u4f4d",
            "\u6578\u91cf",
            "\u55ae\u50f9",
            "\u91d1\u984d",
            "\u5099\u8003",
        ],
        "rows": [],
    }
    assert _classify_table(table) == "estimate"


def test_classify_cost_breakdown_table_as_detail():
    table = {
        "headers": [
            "\ud488\uba85",
            "\uaddc\uaca9",
            "\ub2e8\uc704",
            "\uc218\ub7c9",
            "\uc7ac\ub8cc\ube44_\ub2e8 \uac00",
            "\uae08 \uc561",
            "\ub178\ubb34\ube44_\ub2e8 \uac00",
            "\uae08 \uc561",
            "\uacbd\ube44_\ub2e8 \uac00",
            "\uae08 \uc561",
            "\ud569\uacc4_\ub2e8 \uac00",
            "\uae08 \uc561",
            "\ube44\uace0",
        ],
        "rows": [],
    }
    assert _classify_table(table) == "generic"


def test_classify_spaced_composite_cost_breakdown_table_as_detail():
    table = {
        "headers": [
            "\ud488\uba85",
            "\uaddc\uaca9",
            "\ub2e8\uc704",
            "\uc218\ub7c9",
            "\uc7ac\ub8cc\ube44_\ub2e8 \uac00",
            "\uc7ac\ub8cc\ube44_\uae08 \uc561",
            "\ub178\ubb34\ube44_\ub2e8 \uac00",
            "\ub178\ubb34\ube44_\uae08 \uc561",
            "\uacbd\ube44_\ub2e8 \uac00",
            "\uacbd\ube44_\uae08 \uc561",
            "\ud569\uacc4_\ub2e8 \uac00",
            "\ud569\uacc4_\uae08 \uc561",
            "\ube44\uace0",
        ],
        "rows": [],
    }
    assert _classify_table(table) == "detail"
