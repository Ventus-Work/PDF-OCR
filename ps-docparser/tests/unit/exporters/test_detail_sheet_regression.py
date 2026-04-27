from pathlib import Path

import openpyxl

from exporters.excel_exporter import ExcelExporter


CANONICAL_HEADERS = [
    "품명",
    "규격",
    "단위",
    "수량",
    "재료비_단가",
    "재료비_금액",
    "노무비_단가",
    "노무비_금액",
    "경비_단가",
    "경비_금액",
    "합계_단가",
    "합계_금액",
    "비고",
]


def _detail_section(headers: list[str], row: dict) -> dict:
    return {
        "title": "고려아연 배관 Support 제작_추가_2차분 견적서",
        "tables": [
            {
                "headers": headers,
                "rows": [row],
                "title": "detail",
            }
        ],
    }


def test_detail_sheet_keeps_all_cost_groups_with_canonical_headers(tmp_path: Path):
    row = {
        "품명": "2) SUPPORT 제작",
        "규격": "SUS304",
        "단위": "KG",
        "수량": "117",
        "재료비_단가": "",
        "재료비_금액": "",
        "노무비_단가": "5,160",
        "노무비_금액": "603,720",
        "경비_단가": "146",
        "경비_금액": "17,082",
        "합계_단가": "5,306",
        "합계_금액": "620,802",
        "비고": "",
    }

    out = tmp_path / "detail_canonical.xlsx"
    ExcelExporter().export([_detail_section(CANONICAL_HEADERS, row)], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["내역서"]

    assert [ws.cell(row=2, column=col).value for col in range(1, 14)] == [
        "품명",
        "규격",
        "단위",
        "수량",
        "재료비",
        None,
        "노무비",
        None,
        "경비",
        None,
        "합계",
        None,
        "비고",
    ]
    assert [ws.cell(row=3, column=col).value for col in range(1, 14)] == [
        None,
        None,
        None,
        None,
        "단가",
        "금액",
        "단가",
        "금액",
        "단가",
        "금액",
        "단가",
        "금액",
        None,
    ]
    assert ws.cell(row=4, column=8).value == 603720
    assert ws.cell(row=4, column=10).value == 17082
    assert ws.cell(row=4, column=12).value == 620802


def test_detail_sheet_reads_spaced_composite_alias_headers(tmp_path: Path):
    alias_headers = [
        "품명",
        "규격",
        "단위",
        "수량",
        "재료비_단 가",
        "재료비_금 액",
        "노무비_단 가",
        "노무비_금 액",
        "경비_단 가",
        "경비_금 액",
        "합계_단 가",
        "합계_금 액",
        "비고",
    ]
    row = {
        "품명": "2) SUPPORT 제작",
        "규격": "SUS304",
        "단위": "KG",
        "수량": "117",
        "재료비_단 가": "",
        "재료비_금 액": "",
        "노무비_단 가": "5,160",
        "노무비_금 액": "603,720",
        "경비_단 가": "146",
        "경비_금 액": "17,082",
        "합계_단 가": "5,306",
        "합계_금 액": "620,802",
        "비고": "",
    }

    out = tmp_path / "detail_alias.xlsx"
    ExcelExporter().export([_detail_section(alias_headers, row)], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["내역서"]

    assert ws.cell(row=2, column=5).value == "재료비"
    assert ws.cell(row=2, column=7).value == "노무비"
    assert ws.cell(row=2, column=9).value == "경비"
    assert ws.cell(row=2, column=11).value == "합계"
    assert ws.cell(row=4, column=8).value == 603720
    assert ws.cell(row=4, column=10).value == 17082
    assert ws.cell(row=4, column=12).value == 620802


def test_detail_sheet_reads_headers_with_section_suffix(tmp_path: Path):
    suffix_headers = [
        "품명_1. TT03",
        "규격_1. TT03",
        "단위_1. TT03",
        "수량_1. TT03",
        "재료비_단가_1. TT03",
        "재료비_금액_1. TT03",
        "노무비_단가_1. TT03",
        "노무비_금액_1. TT03",
        "경비_단가_1. TT03",
        "경비_금액_1. TT03",
        "합계_단가_1. TT03",
        "합계_금액_1. TT03",
        "비고_1. TT03",
    ]
    row = {
        "품명_1. TT03": "1) 하지철골 제작 및 설치",
        "규격_1. TT03": "",
        "단위_1. TT03": "TON",
        "수량_1. TT03": "15.75",
        "재료비_단가_1. TT03": "4,300,000",
        "재료비_금액_1. TT03": "67,725,000",
        "노무비_단가_1. TT03": "3,800,000",
        "노무비_금액_1. TT03": "59,850,000",
        "경비_단가_1. TT03": "146,000",
        "경비_금액_1. TT03": "2,299,500",
        "합계_단가_1. TT03": "8,246,000",
        "합계_금액_1. TT03": "129,874,500",
        "비고_1. TT03": "아연도금",
    }

    out = tmp_path / "detail_suffix.xlsx"
    ExcelExporter().export([_detail_section(suffix_headers, row)], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["내역서"]

    assert ws.cell(row=4, column=1).value == "1) 하지철골 제작 및 설치"
    assert ws.cell(row=4, column=6).value == 67725000
    assert ws.cell(row=4, column=8).value == 59850000
    assert ws.cell(row=4, column=12).value == 129874500
