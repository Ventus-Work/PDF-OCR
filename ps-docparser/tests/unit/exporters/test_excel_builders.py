"""tests/unit/exporters/test_excel_builders.py — Phase 12.5 신규 빌더 단위 테스트"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from exporters.excel_builders import (
    _build_text_sheet,
    _build_notes_sheet,
    _build_conditions_sheet,
    _build_crossref_sheet,
    _build_meta_sheet,
    _build_drawing_meta_sheet,
)


# ─────────────────────────────────────────────────────────────
# 공통 픽스처
# ─────────────────────────────────────────────────────────────

@pytest.fixture
def pumsem_rich_sections():
    return [
        {
            "section_id": "1-1-1",
            "title": "일반사항",
            "department": "공통부문",
            "chapter": "제1장 총칙",
            "page": 5,
            "clean_text": "본 품셈은 표준 단위 공사량 산정 기준이다.",
            "tables": [],
            "notes": ["본 품셈은 2024년 기준이다", "야간작업 시 25% 가산"],
            "conditions": [
                {"type": "가산", "condition": "야간작업 시", "rate": "25%"},
                {"type": "감산", "condition": "건설기계 사용 시", "rate": "10%"},
            ],
            "cross_references": [
                {"target_section_id": "2-1-3", "target_chapter": "제2장",
                 "context": "제2장 2-1-3 준용"},
            ],
            "revision_year": "2024",
            "unit_basis": "m³당",
        },
    ]


@pytest.fixture
def bom_minimal_sections():
    return [
        {
            "section_id": "BOM-1",
            "title": "BILL OF MATERIALS #1",
            "department": None,
            "chapter": None,
            "page": 1,
            "clean_text": "",
            "tables": [],
            "notes": [],
            "conditions": [],
            "cross_references": [],
            "revision_year": None,
            "unit_basis": None,
        },
    ]


def _ws():
    wb = Workbook()
    wb.remove(wb.active)
    return wb.create_sheet("test")


# ─────────────────────────────────────────────────────────────
# Step 12.5-1: _build_text_sheet
# ─────────────────────────────────────────────────────────────

class TestBuildTextSheet:
    def test_single_section_with_text(self, pumsem_rich_sections):
        ws = _ws()
        _build_text_sheet(ws, pumsem_rich_sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2  # 헤더 + 1 데이터 행

    def test_multi_section_header_order(self, pumsem_rich_sections):
        ws = _ws()
        _build_text_sheet(ws, pumsem_rich_sections)
        headers = [c.value for c in ws[1]]
        assert headers == ["섹션 ID", "부문", "장", "제목", "페이지", "본문"]

    def test_long_text_wraps(self, pumsem_rich_sections):
        """500자 본문 → F열 wrap_text=True 유지 (_ALIGN_LEFT 기본값)"""
        sections = [dict(pumsem_rich_sections[0])]
        sections[0]["clean_text"] = "가" * 500
        ws = _ws()
        _build_text_sheet(ws, sections)
        f_cell = ws.cell(row=2, column=6)
        assert f_cell.alignment.wrap_text is True

    def test_empty_clean_text_skipped(self, bom_minimal_sections):
        ws = _ws()
        _build_text_sheet(ws, bom_minimal_sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1  # 헤더만


# ─────────────────────────────────────────────────────────────
# Step 12.5-2: _build_notes_sheet
# ─────────────────────────────────────────────────────────────

class TestBuildNotesSheet:
    def test_single_note(self, pumsem_rich_sections):
        sections = [dict(pumsem_rich_sections[0], notes=["단일 주석"])]
        ws = _ws()
        _build_notes_sheet(ws, sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2

    def test_multi_notes_flatten(self):
        sections = [
            {"section_id": "A", "title": "A절", "page": 1,
             "notes": ["A주석1", "A주석2"]},
            {"section_id": "B", "title": "B절", "page": 2,
             "notes": ["B주석1", "B주석2", "B주석3"]},
        ]
        ws = _ws()
        _build_notes_sheet(ws, sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 6  # 헤더 + 5

    def test_note_number_per_section(self):
        """섹션 내 주석 번호는 섹션별로 1부터 재시작."""
        sections = [
            {"section_id": "A", "title": "A", "page": 1, "notes": ["n1", "n2"]},
            {"section_id": "B", "title": "B", "page": 2, "notes": ["n1"]},
        ]
        ws = _ws()
        _build_notes_sheet(ws, sections)
        note_nums = [ws.cell(row=r, column=4).value for r in range(2, 5)]
        assert note_nums == [1, 2, 1]


# ─────────────────────────────────────────────────────────────
# Step 12.5-3: _build_conditions_sheet
# ─────────────────────────────────────────────────────────────

class TestBuildConditionsSheet:
    def test_single_condition(self, pumsem_rich_sections):
        sections = [dict(pumsem_rich_sections[0],
                         conditions=[{"type": "가산", "condition": "야간", "rate": "25%"}])]
        ws = _ws()
        _build_conditions_sheet(ws, sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2

    def test_multiple_types(self, pumsem_rich_sections):
        ws = _ws()
        _build_conditions_sheet(ws, pumsem_rich_sections)
        types = [ws.cell(row=r, column=4).value for r in range(2, 4)]
        assert "가산" in types
        assert "감산" in types

    def test_rate_preserved(self):
        sections = [{"section_id": "X", "title": "X", "page": 1,
                     "conditions": [{"type": "할증", "condition": "한냉지", "rate": "30%"}]}]
        ws = _ws()
        _build_conditions_sheet(ws, sections)
        assert ws.cell(row=2, column=6).value == "30%"


# ─────────────────────────────────────────────────────────────
# Step 12.5-4: _build_crossref_sheet
# ─────────────────────────────────────────────────────────────

class TestBuildCrossrefSheet:
    def test_single_crossref(self, pumsem_rich_sections):
        ws = _ws()
        _build_crossref_sheet(ws, pumsem_rich_sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2

    def test_target_chapter_empty(self):
        sections = [{"section_id": "A", "title": "A", "page": 1,
                     "cross_references": [
                         {"target_section_id": "3-1", "target_chapter": "", "context": "준용"}
                     ]}]
        ws = _ws()
        _build_crossref_sheet(ws, sections)
        assert ws.cell(row=2, column=4).value == ""

    def test_multi_refs_per_section(self):
        sections = [{"section_id": "A", "title": "A", "page": 1,
                     "cross_references": [
                         {"target_section_id": "1", "target_chapter": "제1장", "context": "ctx1"},
                         {"target_section_id": "2", "target_chapter": "제2장", "context": "ctx2"},
                         {"target_section_id": "3", "target_chapter": "제3장", "context": "ctx3"},
                     ]}]
        ws = _ws()
        _build_crossref_sheet(ws, sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 4
        assert ws.cell(row=2, column=1).value == "A"
        assert ws.cell(row=4, column=1).value == "A"


# ─────────────────────────────────────────────────────────────
# Step 12.5-5: _build_meta_sheet
# ─────────────────────────────────────────────────────────────

class TestBuildMetaSheet:
    def test_revision_year_only(self):
        sections = [{"section_id": "A", "title": "A", "page": 1,
                     "revision_year": "2024", "unit_basis": ""}]
        ws = _ws()
        _build_meta_sheet(ws, sections)
        assert ws.cell(row=2, column=4).value == "2024"
        assert ws.cell(row=2, column=5).value == ""

    def test_unit_basis_only(self):
        sections = [{"section_id": "B", "title": "B", "page": 2,
                     "revision_year": "", "unit_basis": "m²당"}]
        ws = _ws()
        _build_meta_sheet(ws, sections)
        assert ws.cell(row=2, column=4).value == ""
        assert ws.cell(row=2, column=5).value == "m²당"

    def test_both_empty_row_skipped(self):
        sections = [{"section_id": "C", "title": "C", "page": 3,
                     "revision_year": "", "unit_basis": ""}]
        ws = _ws()
        _build_meta_sheet(ws, sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1  # 헤더만

    def test_none_values_normalized(self, bom_minimal_sections):
        """BOM 모드 None → "" 정규화 후 행 스킵."""
        ws = _ws()
        _build_meta_sheet(ws, bom_minimal_sections)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1


# ─────────────────────────────────────────────────────────────
# Phase 14: _build_drawing_meta_sheet
# ─────────────────────────────────────────────────────────────

class TestBuildDrawingMetaSheet:
    def test_empty_dict_skipped(self):
        ws = _ws()
        _build_drawing_meta_sheet(ws, {})
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 0

    def test_none_values_skipped(self):
        ws = _ws()
        _build_drawing_meta_sheet(ws, {"dwg_no": None, "rev": None})
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2  # 제목행, 헤더행 (데이터 없음)

    def test_values_rendered(self):
        ws = _ws()
        data = {
            "dwg_no": "ABC-123",
            "rev": "0",
            "title": "SUPPORT PLAN",
            "unknown_field": "test"
        }
        _build_drawing_meta_sheet(ws, data)
        rows = list(ws.iter_rows(values_only=True))
        
        # 필드 순서대로 출력되어야 하므로: 제목(1) + 헤더(1) + dwg_no(1) + rev(1) + title(1)
        # unknown_field는 _DRAWING_META_FIELD_ORDER에 없으므로 무시됨
        assert len(rows) == 5
        assert ws.cell(row=3, column=2).value == "ABC-123"
        assert ws.cell(row=4, column=2).value == "0"
        assert ws.cell(row=5, column=2).value == "SUPPORT PLAN"
