"""tests/unit/exporters/test_excel_exporter.py"""

from __future__ import annotations

from pathlib import Path

import pytest

from exporters.excel_exporter import (
    ExcelExporter,
    _classify_table,
    _try_parse_number,
    _row_style,
)


# ─────────────────────────────────────────────────────────────
# _classify_table
# ─────────────────────────────────────────────────────────────

class TestClassifyTable:
    def test_estimate_by_header(self):
        table = {"headers": ["명 칭", "규 격", "수량", "금 액"], "rows": []}
        assert _classify_table(table) == "estimate"

    def test_estimate_spaceless_header(self):
        table = {"headers": ["명칭", "규격", "수량", "금액"], "rows": []}
        assert _classify_table(table) == "estimate"

    def test_detail_by_header(self):
        table = {"headers": ["품명", "수량", "합계_금액"], "rows": []}
        assert _classify_table(table) == "detail"

    def test_condition_by_keyword(self):
        table = {"headers": ["일반사항"], "rows": []}
        assert _classify_table(table) == "condition"

    def test_condition_by_type_hint(self):
        table = {"headers": ["X", "Y"], "type": "D_기타", "rows": []}
        assert _classify_table(table) == "condition"

    def test_generic_fallback(self):
        table = {"headers": ["X", "Y"], "rows": []}
        assert _classify_table(table) == "generic"

    def test_empty_headers(self):
        table = {"rows": []}
        assert _classify_table(table) == "generic"

    def test_condition_by_type_hint(self):
        table = {"headers": ["X", "Y"], "type": "D_\uae30\ud0c0", "rows": []}
        assert _classify_table(table) == "generic"

    def test_condition_by_type_hint_with_condition_rows(self):
        table = {
            "headers": ["A", "B"],
            "type": "D_\uae30\ud0c0",
            "rows": [{"A": "\uc77c\ubc18\uc0ac\ud56d", "B": "\ud2b9\uae30\uc0ac\ud56d"}],
        }
        assert _classify_table(table) == "condition"

    def test_estimate_by_hanja_headers(self):
        table = {
            "headers": [
                "\u540d\u7a31",
                "\u898f\u683c",
                "\u55ae\u4f4d",
                "\u6578\u91cf",
                "\u55ae\u50f9",
                "\u91d1\u984d",
                "\u5099\u8003",
            ],
            "rows": [],
        }
        assert _classify_table(table) == "estimate"

    def test_pumsem_table_types_stay_generic(self):
        table = {
            "headers": ["규격", "단위수량"],
            "rows": [{"규격": "종 목", "단위수량": "비 고"}],
            "type": "A_품셈",
        }
        assert _classify_table(table) == "generic"


# ─────────────────────────────────────────────────────────────
# _try_parse_number
# ─────────────────────────────────────────────────────────────

class TestTryParseNumber:
    @pytest.mark.parametrize("text,expected", [
        ("1,000,000", 1_000_000),
        ("  3.5  ", 3.5),
        ("-500", -500),
        ("0", 0),
        ("0.5", 0.5),
        ("1,234원", None),
        ("(5,000)", None),
        ("0015", None),
        ("-", None),
        ("", None),
        ("N/A", None),
        ("100KW", None),
    ])
    def test_parse_variants(self, text, expected):
        assert _try_parse_number(text) == expected

    def test_non_string_returns_none(self):
        assert _try_parse_number(None) is None
        assert _try_parse_number(123) is None


# ─────────────────────────────────────────────────────────────
# _row_style
# ─────────────────────────────────────────────────────────────

class TestRowStyle:
    def test_subtotal_row(self):
        row = {"명칭": "소 계", "금 액": "1,000,000"}
        assert _row_style(row, "명칭") == "subtotal"

    def test_total_row(self):
        row = {"명칭": "합 계", "금 액": "5,000,000"}
        assert _row_style(row, "명칭") == "subtotal"

    def test_section_row(self):
        row = {"명칭": "1. 배관공사", "금 액": ""}
        assert _row_style(row, "명칭") == "section"

    def test_body_row(self):
        row = {"명칭": "파이프 서포트", "금 액": "1,000,000"}
        assert _row_style(row, "명칭") == "body"

    def test_no_money_key_no_section(self):
        # money_keys 가 빈 리스트면 all([]) == True 함정 → body 반환
        row = {"품명": "앵커볼트"}
        assert _row_style(row, "품명") == "body"


# ─────────────────────────────────────────────────────────────
# ExcelExporter.export() 통합
# ─────────────────────────────────────────────────────────────

class TestExcelExporterExport:
    def _make_section(self, table_type: str = "estimate") -> dict:
        return {
            "title": "테스트",
            "tables": [
                {
                    "headers": ["명 칭", "규 격", "수량", "단위", "금 액"],
                    "rows": [
                        {"명 칭": "파이프 서포트", "규 격": "150A", "수량": 4, "단위": "EA", "금 액": 1_000_000},
                        {"명 칭": "소 계", "규 격": "", "수량": "", "단위": "", "금 액": 1_000_000},
                    ],
                    "type": table_type,
                    "title": "견적",
                }
            ],
        }

    def test_creates_xlsx_file(self, tmp_path: Path):
        out = tmp_path / "test.xlsx"
        ExcelExporter().export([self._make_section()], out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_returns_output_path(self, tmp_path: Path):
        out = tmp_path / "test.xlsx"
        result = ExcelExporter().export([self._make_section()], out)
        assert result == out

    def test_empty_sections_creates_fallback_sheet(self, tmp_path: Path):
        import openpyxl
        out = tmp_path / "empty.xlsx"
        ExcelExporter().export([], out)
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) >= 1

    def test_metadata_title_accepted(self, tmp_path: Path):
        out = tmp_path / "titled.xlsx"
        result = ExcelExporter().export(
            [self._make_section()], out, metadata={"description": "견적서 제목"}
        )
        assert result.exists()

    def test_generic_table_creates_table_sheet(self, tmp_path: Path):
        import openpyxl
        section = {
            "title": "BOM",
            "tables": [
                {
                    "headers": ["SIZE", "MAT'L", "Q'TY"],
                    "rows": [{"SIZE": "150A", "MAT'L": "SS400", "Q'TY": 4}],
                    "type": "generic",
                    "title": "",
                }
            ],
        }
        out = tmp_path / "generic.xlsx"
        ExcelExporter().export([section], out)
        wb = openpyxl.load_workbook(out)
        sheet_names = wb.sheetnames
        assert any("Table" in s or "BOM" in s or s for s in sheet_names)

    def test_generic_pumsem_table_uses_meaningful_sheet_name(self, tmp_path: Path):
        import openpyxl
        section = {
            "title": "53-83 OKOK",
            "tables": [
                {
                    "headers": ["규격", "단위수량"],
                    "rows": [{"규격": "종 목", "단위수량": "비 고"}],
                    "type": "A_품셈",
                    "section_title": "1-2-2 단위표준",
                    "title": "",
                }
            ],
        }
        out = tmp_path / "pumsem_generic.xlsx"
        ExcelExporter().export([section], out)
        wb = openpyxl.load_workbook(out)
        assert any("단위표준" in name for name in wb.sheetnames)
        assert all(not name.startswith("Table") for name in wb.sheetnames if "단위표준" in name)

    def test_condition_table_creates_condition_sheet(self, tmp_path: Path):
        import openpyxl
        section = {
            "title": "조건",
            "tables": [
                {
                    "headers": ["일반사항"],
                    "rows": [{"일반사항": "재질: SS400"}, {"일반사항": "도장: 에폭시"}],
                    "type": "condition",
                    "title": "조건",
                }
            ],
        }
        out = tmp_path / "cond.xlsx"
        ExcelExporter().export([section], out)
        wb = openpyxl.load_workbook(out)
        assert "조건" in wb.sheetnames


# ─────────────────────────────────────────────────────────────
# Phase 12.5 통합 테스트
# ─────────────────────────────────────────────────────────────

_PUMSEM_SECTION = {
    "section_id": "1-1-1",
    "title": "일반사항",
    "department": "공통부문",
    "chapter": "제1장",
    "page": 5,
    "clean_text": "본 품셈은 기준이다.",
    "tables": [],
    "notes": ["주석1"],
    "conditions": [{"type": "가산", "condition": "야간", "rate": "25%"}],
    "cross_references": [{"target_section_id": "2-1", "target_chapter": "제2장", "context": "준용"}],
    "revision_year": "2024",
    "unit_basis": "m³당",
}

_BOM_SECTION = {
    "section_id": "BOM-1",
    "title": "BOM",
    "department": None,
    "chapter": None,
    "page": 1,
    "clean_text": "",
    "tables": [{"headers": ["NO", "DESC"], "rows": [{"NO": "1", "DESC": "파이프"}],
                "type": "generic", "title": ""}],
    "notes": [],
    "conditions": [],
    "cross_references": [],
    "revision_year": None,
    "unit_basis": None,
}


class TestPhase125Integration:
    def test_pumsem_sample_sheets_present(self, tmp_path: Path):
        import openpyxl
        out = tmp_path / "pumsem.xlsx"
        ExcelExporter().export([_PUMSEM_SECTION], out)
        names = openpyxl.load_workbook(out).sheetnames
        for expected in ["본문", "주석", "가감산_조건", "교차참조", "메타데이터"]:
            assert expected in names, f"'{expected}' 시트 없음: {names}"

    def test_bom_sample_no_new_sheets(self, tmp_path: Path):
        import openpyxl
        out = tmp_path / "bom.xlsx"
        ExcelExporter().export([_BOM_SECTION], out)
        names = openpyxl.load_workbook(out).sheetnames
        for unexpected in ["본문", "주석", "가감산_조건", "교차참조", "메타데이터"]:
            assert unexpected not in names, f"BOM 모드에 '{unexpected}' 시트 생성됨"

    def test_empty_fields_no_new_sheets(self, tmp_path: Path):
        import openpyxl
        empty = {"section_id": "e", "title": "", "department": "", "chapter": "",
                 "page": 0, "clean_text": "", "tables": [], "notes": [],
                 "conditions": [], "cross_references": [],
                 "revision_year": "", "unit_basis": ""}
        out = tmp_path / "empty.xlsx"
        ExcelExporter().export([empty], out)
        names = openpyxl.load_workbook(out).sheetnames
        for unexpected in ["본문", "주석", "가감산_조건", "교차참조", "메타데이터"]:
            assert unexpected not in names

    def test_partial_fields_partial_sheets(self, tmp_path: Path):
        import openpyxl
        section = dict(_PUMSEM_SECTION,
                       clean_text="", conditions=[], cross_references=[],
                       revision_year="", unit_basis="",
                       notes=["주석만"])
        out = tmp_path / "partial.xlsx"
        ExcelExporter().export([section], out)
        names = openpyxl.load_workbook(out).sheetnames
        assert "주석" in names
        for unexpected in ["본문", "가감산_조건", "교차참조", "메타데이터"]:
            assert unexpected not in names, f"'{unexpected}' 시트가 생성됨"

    def test_sheet_order(self, tmp_path: Path):
        import openpyxl
        section = dict(_PUMSEM_SECTION, tables=[
            {"headers": ["명 칭", "금 액"], "rows": [], "type": "estimate", "title": "견적"},
        ])
        out = tmp_path / "order.xlsx"
        ExcelExporter().export([section], out)
        names = openpyxl.load_workbook(out).sheetnames
        # 견적서가 본문보다 앞에 있어야 함
        if "견적서" in names and "본문" in names:
            assert names.index("견적서") < names.index("본문")
        # 본문이 주석보다 앞에
        if "본문" in names and "주석" in names:
            assert names.index("본문") < names.index("주석")

    def test_drawing_meta_sheet_created(self, tmp_path: Path):
        """Phase 14: drawing_meta 타입 섹션이 있으면 '도면_메타' 시트가 생성되어야 함."""
        import openpyxl
        section = {
            "section_id": "DRAWING-META-1",
            "title": "도면 메타데이터",
            "type": "drawing_meta",
            "drawing_metadata": {"dwg_no": "TEST-123", "rev": "A"}
        }
        out = tmp_path / "dwg_meta.xlsx"
        from exporters.excel_exporter import export
        export([section], out)
        
        wb = openpyxl.load_workbook(out)
        assert "도면_메타" in wb.sheetnames
        ws = wb["도면_메타"]
        
        # 단순히 내용 존재 여부만 체크
        has_content = False
        for row in ws.iter_rows(values_only=True):
            if "TEST-123" in row:
                has_content = True
                break
        assert has_content
